import base64
import csv
import functools
import io
import json
import os
import platform as _platform_module
import shutil
import urllib.error
import urllib.request
from types import SimpleNamespace

from flask import (
    Flask, flash, redirect, render_template,
    request, url_for, send_file, Response, session, jsonify,
)
from werkzeug.security import generate_password_hash, check_password_hash
from apscheduler.schedulers.background import BackgroundScheduler
import openpyxl

import database as db

APP_VERSION = "1.1.0"
GITHUB_REPO = "MidnightMight/Job_Application_tracker"

# Limits used in API helpers.
_MAX_JOB_DESC_LENGTH  = 4000   # characters sent to the LLM
_MAX_ERROR_MSG_LENGTH = 120    # characters of raw error text exposed to clients
_MAX_PROFILE_LENGTH   = 8000   # characters stored for user profile summary from PDF


def _detect_deployment_mode() -> str:
    """Return 'docker' (full features) or 'local' (single-user, no AI).

    Precedence:
    1. DEPLOYMENT_MODE env var (explicit override)
    2. Presence of /.dockerenv (running inside a container)
    3. DB_PATH env var set (Docker Compose / container typically sets this)
    4. OS: Windows or macOS → 'local'; Linux/other → 'docker'
    """
    explicit = os.environ.get("DEPLOYMENT_MODE", "").lower()
    if explicit in ("docker", "local"):
        return explicit
    if os.path.exists("/.dockerenv"):
        return "docker"
    if os.environ.get("DB_PATH"):
        return "docker"
    system = _platform_module.system()
    if system in ("Windows", "Darwin"):
        return "local"
    return "docker"


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "job-tracker-secret-key-change-me")

db.init_db()

# Compute once at startup; available everywhere as DEPLOYMENT_MODE.
DEPLOYMENT_MODE: str = _detect_deployment_mode()

app.jinja_env.globals["enumerate"] = enumerate


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def login_required(f):
    """Decorator: redirect to login if auth is enabled and user is not logged in."""
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if db.get_setting("login_enabled", "0") == "1":
            if not session.get("user_id"):
                return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Background scheduler — reminder checks
# ---------------------------------------------------------------------------

def _check_and_create_reminders():
    """Scheduled task: create inbox reminders for long-pending applications."""
    try:
        if db.get_setting("reminder_enabled", "1") != "1":
            return
        days = int(db.get_setting("reminder_days", "3"))
        for app_record in db.get_pending_for_reminders(days):
            msg = (
                f"'{app_record['job_desc'] or 'Application'}' at {app_record['company']} "
                f"has been pending ({app_record['status'].replace('_', ' ')}) "
                f"for {app_record['duration']} days."
            )
            db.create_reminder(app_record["id"], msg)
    except Exception:
        pass  # Never crash the scheduler thread


# Start scheduler only once (skip in Werkzeug reloader child process).
if os.environ.get("WERKZEUG_RUN_MAIN") != "false":
    _scheduler = BackgroundScheduler(daemon=True)
    _scheduler.add_job(_check_and_create_reminders, "interval", hours=1, id="reminders")
    _scheduler.start()
    # Also run once on startup so the inbox is populated immediately.
    _check_and_create_reminders()


@app.context_processor
def inject_globals():
    from datetime import date
    _profile_complete = bool(
        db.get_setting("user_profile_skills", "").strip()
        or db.get_setting("user_profile_experience", "").strip()
        or db.get_setting("user_profile_summary", "").strip()
    )
    return {
        "years": db.YEARS,
        "current_year_for_footer": date.today().year,
        "unread_reminder_count": db.get_unread_reminder_count(),
        "login_enabled": db.get_setting("login_enabled", "0") == "1",
        "current_user": session.get("username"),
        "app_version": APP_VERSION,
        "ollama_enabled": db.get_setting("ollama_enabled", "0") == "1",
        "ai_fit_enabled": db.get_setting("ai_fit_enabled", "0") == "1",
        "user_profile_complete": _profile_complete,
        "deployment_mode": DEPLOYMENT_MODE,
    }


# ---------------------------------------------------------------------------
# Onboarding (first-run setup)
# ---------------------------------------------------------------------------

# Endpoints that must remain accessible before onboarding is complete.
_ONBOARDING_EXEMPT = {"onboarding", "static", "login", "logout"}


@app.before_request
def _check_onboarding():
    """Redirect every request to /onboarding until the first-run wizard is done."""
    if request.endpoint in _ONBOARDING_EXEMPT or request.endpoint is None:
        return
    if db.get_setting("onboarding_complete", "0") == "0":
        return redirect(url_for("onboarding"))


@app.route("/onboarding", methods=["GET", "POST"])
def onboarding():
    # Already completed — go to dashboard.
    if db.get_setting("onboarding_complete", "0") == "1":
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "setup_admin":
            # Validate and create the admin user, then enable login.
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            password2 = request.form.get("password2", "")
            clear_demo = request.form.get("clear_demo") == "1"

            errors = []
            if not username:
                errors.append("Username is required.")
            if not password:
                errors.append("Password is required.")
            elif len(password) < 8:
                errors.append("Password must be at least 8 characters.")
            elif password != password2:
                errors.append("Passwords do not match.")

            if errors:
                for e in errors:
                    flash(e, "danger")
                return render_template("onboarding.html", step="admin",
                                       deployment_mode=DEPLOYMENT_MODE)

            pw_hash = generate_password_hash(password)
            ok, msg = db.add_user(username, pw_hash, is_admin=True)
            if not ok:
                flash(msg, "danger")
                return render_template("onboarding.html", step="admin",
                                       deployment_mode=DEPLOYMENT_MODE)

            db.set_setting("login_enabled", "1")
            session["user_id"] = db.get_user_by_username(username)["id"]
            session["username"] = username

            if clear_demo:
                db.clear_demo_data()

            db.set_setting("onboarding_complete", "1")
            flash(f"Welcome, {username}! Your account has been created and login is enabled.", "success")
            return redirect(url_for("dashboard"))

        elif action == "skip":
            # No admin account; login stays disabled.
            clear_demo = request.form.get("clear_demo") == "1"
            if clear_demo:
                db.clear_demo_data()
            db.set_setting("onboarding_complete", "1")
            flash("Welcome to Job Tracker! You can set up login any time in Settings.", "info")
            return redirect(url_for("dashboard"))

        # Unknown action — fall through to render.

    # Advance to step 2 when the Continue button passes ?next_step=1.
    if request.args.get("next_step") == "1":
        return render_template("onboarding.html", step="admin",
                               deployment_mode=DEPLOYMENT_MODE)

    return render_template("onboarding.html", step="welcome",
                           deployment_mode=DEPLOYMENT_MODE)


# ---------------------------------------------------------------------------
# Login / Logout
# ---------------------------------------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if db.get_setting("login_enabled", "0") != "1":
        return redirect(url_for("dashboard"))
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = db.get_user_by_username(username)
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            flash(f"Welcome back, {user['username']}!", "success")
            # Only allow redirects to internal paths: must start with '/'
            # and contain no scheme or netloc, preventing open-redirect attacks.
            raw_next = request.form.get("next", "")
            from urllib.parse import urlparse as _urlparse
            _p = _urlparse(raw_next)
            _safe = (
                raw_next
                and raw_next.startswith("/")
                and not raw_next.startswith("//")
                and not _p.netloc
                and not _p.scheme
            )
            next_url = raw_next if _safe else url_for("dashboard")
            return redirect(next_url)
        flash("Invalid username or password.", "danger")
    return render_template("login.html", next=request.args.get("next", ""))


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@app.route("/")
@login_required
def dashboard():
    current_year = 2025
    stats = db.get_stats(year=current_year)
    status_counts = db.get_status_counts(year=current_year)
    apps_per_year = db.get_apps_per_year()
    success_per_year = db.get_success_rate_per_year()
    keyword_freq = db.get_company_note_frequency()
    return render_template(
        "dashboard.html",
        stats=stats,
        status_counts=status_counts,
        apps_per_year=apps_per_year,
        success_per_year=success_per_year,
        keyword_freq=keyword_freq,
        current_year=current_year,
        years=db.YEARS,
    )


# ---------------------------------------------------------------------------
# Global search
# ---------------------------------------------------------------------------

@app.route("/search")
@login_required
def search():
    query = request.args.get("q", "").strip()
    results = []
    if len(query) >= 2:
        results = db.search_applications(query)
    return render_template("search.html", query=query, results=results)


# ---------------------------------------------------------------------------
# Year view
# ---------------------------------------------------------------------------

@app.route("/year/<int:year>")
def year_view(year):
    status_filter = request.args.get("status", "")
    apps = db.get_applications(
        year=year,
        status=status_filter if status_filter else None,
    )
    stats = db.get_stats(year=year)
    return render_template(
        "year_view.html",
        apps=apps,
        year=year,
        stats=stats,
        years=db.YEARS,
        status_options=db.get_status_options(),
        selected_status=status_filter,
    )


# ---------------------------------------------------------------------------
# Application detail
# ---------------------------------------------------------------------------

@app.route("/application/<int:app_id>")
def application_detail(app_id):
    application = db.get_application(app_id)
    if not application:
        flash("Application not found.", "danger")
        return redirect(url_for("dashboard"))
    timeline = db.get_application_timeline(app_id)
    return render_template(
        "application_detail.html",
        app=application,
        timeline=timeline,
    )


# ---------------------------------------------------------------------------
# Add / edit / delete application
# ---------------------------------------------------------------------------

@app.route("/application/add", methods=["GET", "POST"])
def add_application():
    if request.method == "POST":
        company     = request.form.get("company", "").strip()
        job_desc    = request.form.get("job_desc", "").strip()
        date_applied = request.form.get("date_applied", "").strip()

        # Check for duplicates unless the user has explicitly confirmed.
        if request.form.get("force_add") != "1":
            duplicates = db.find_duplicate_applications(company, job_desc, date_applied)
            if duplicates:
                # Re-render the form with the submitted values preserved and
                # trigger the duplicate-warning modal.
                form_data = SimpleNamespace(
                    job_desc=job_desc,
                    company=company,
                    team=request.form.get("team", ""),
                    contact=request.form.get("contact", ""),
                    date_applied=date_applied,
                    status=request.form.get("status", "Select_Status"),
                    success_chance=request.form.get("success_chance", "0"),
                    link=request.form.get("link", ""),
                    cover_letter=1 if request.form.get("cover_letter") else 0,
                    resume=1 if request.form.get("resume") else 0,
                    comment=request.form.get("comment", ""),
                    additional_notes=request.form.get("additional_notes", ""),
                )
                return render_template(
                    "application_form.html",
                    app=form_data,
                    companies=db.get_companies(),
                    status_options=db.get_status_options(),
                    action="Add",
                    duplicate_warning=True,
                    duplicates=duplicates,
                )

        db.add_application(request.form)
        flash("Application added.", "success")
        year = request.form.get("date_applied", "")[:4] or "2025"
        return redirect(url_for("year_view", year=year))
    companies_list = db.get_companies()
    return render_template(
        "application_form.html",
        app=None,
        companies=companies_list,
        status_options=db.get_status_options(),
        action="Add",
    )


@app.route("/application/edit/<int:app_id>", methods=["GET", "POST"])
def edit_application(app_id):
    application = db.get_application(app_id)
    if not application:
        flash("Application not found.", "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        db.update_application(app_id, request.form)
        flash("Application updated.", "success")
        year = request.form.get("date_applied", "")[:4] or "2025"
        return redirect(url_for("year_view", year=year))
    companies_list = db.get_companies()
    return render_template(
        "application_form.html",
        app=application,
        companies=companies_list,
        status_options=db.get_status_options(),
        action="Edit",
    )


@app.route("/application/delete/<int:app_id>", methods=["POST"])
def delete_application(app_id):
    application = db.get_application(app_id)
    year = application["date_applied"][:4] if application else "2025"
    db.delete_application(app_id)
    flash("Application deleted.", "warning")
    return redirect(url_for("year_view", year=year))


# ---------------------------------------------------------------------------
# Bulk actions (multi-select in year view)
# ---------------------------------------------------------------------------

@app.route("/applications/bulk-action", methods=["POST"])
def bulk_action():
    """Handle bulk operations (delete / set-field) on multiple applications."""
    action      = request.form.get("action", "")
    year        = request.form.get("year",   "2025")
    status_filter = request.form.get("status_filter", "")

    raw_ids = request.form.getlist("selected_ids")
    try:
        selected_ids = [int(x) for x in raw_ids if str(x).isdigit()]
    except ValueError:
        selected_ids = []

    redirect_kwargs: dict = {"year": year}
    if status_filter:
        redirect_kwargs["status"] = status_filter

    if not selected_ids:
        flash("No applications selected.", "warning")
        return redirect(url_for("year_view", **redirect_kwargs))

    if action == "delete":
        count = db.bulk_delete_applications(selected_ids)
        flash(f"Deleted {count} application(s).", "warning")

    elif action == "set_status":
        new_status = request.form.get("bulk_status", "").strip()
        if not new_status:
            flash("Please choose a status.", "warning")
        else:
            count = db.bulk_update_applications(selected_ids, "status", new_status)
            flash(
                f"Status set to '{new_status.replace('_', ' ')}' "
                f"for {count} application(s).",
                "success",
            )

    elif action == "set_date_applied":
        new_date = request.form.get("bulk_date_applied", "").strip()
        if not new_date:
            flash("Please enter a date.", "warning")
        else:
            count = db.bulk_update_applications(selected_ids, "date_applied", new_date)
            flash(f"Date Applied set to {new_date} for {count} application(s).", "success")

    elif action == "set_last_contact":
        new_date = request.form.get("bulk_last_contact", "").strip()
        if not new_date:
            flash("Please enter a date.", "warning")
        else:
            count = db.bulk_update_applications(selected_ids, "last_contact_date", new_date)
            flash(f"Last Contact set to {new_date} for {count} application(s).", "success")

    elif action == "set_cover_letter":
        value = 1 if request.form.get("bulk_cover_letter") == "1" else 0
        label = "Yes" if value else "No"
        count = db.bulk_update_applications(selected_ids, "cover_letter", value)
        flash(f"Cover Letter set to {label} for {count} application(s).", "success")

    elif action == "set_resume":
        value = 1 if request.form.get("bulk_resume") == "1" else 0
        label = "Yes" if value else "No"
        count = db.bulk_update_applications(selected_ids, "resume", value)
        flash(f"Resume set to {label} for {count} application(s).", "success")

    else:
        flash("Unknown bulk action.", "warning")

    return redirect(url_for("year_view", **redirect_kwargs))


# ---------------------------------------------------------------------------
# CSV / Excel bulk import  (Mouser/DigiKey-style column mapping)
# ---------------------------------------------------------------------------

# Recognised Excel file extensions.
_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}


def _excel_sheet_to_csv(workbook_bytes: bytes, sheet_name: str, start_row: int = 1) -> str:
    """Read one sheet from an Excel workbook and return it as a CSV string.

    ``start_row`` is 1-indexed; rows before it are skipped so the first row
    of the returned CSV is the header row chosen by the user.
    """
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name]
    out = io.StringIO()
    writer = csv.writer(out)
    for row in ws.iter_rows(min_row=max(1, start_row), values_only=True):
        writer.writerow([("" if cell is None else str(cell)) for cell in row])
    wb.close()
    return out.getvalue()


# Application fields that can be mapped from a CSV column.
CSV_IMPORT_FIELDS = [
    ("",                "(ignore this column)"),
    ("company",         "Company *"),
    ("date_applied",    "Date Applied *"),
    ("job_desc",        "Job Description / Role"),
    ("team",            "Team / Division"),
    ("status",          "Status"),
    ("cover_letter",    "Cover Letter (1/0)"),
    ("resume",          "Resume (1/0)"),
    ("success_chance",  "Success Chance"),
    ("link",            "Application Link"),
    ("contact",         "Known Contact / Connection"),
    ("comment",         "Comment / Notes"),
    ("additional_notes","Additional Notes"),
]


@app.route("/application/import", methods=["GET", "POST"])
def import_csv():
    """
    Step 1  (GET or first POST with a file): Upload file.
              – CSV  → proceed straight to column mapping.
              – Excel → show sheet picker.
    Step 1b (POST with stage='sheet'): Excel sheet selected → convert to CSV
              and proceed to column mapping.
    Step 2  (POST with stage='import'): Run the import and show results.
    """
    if request.method == "GET":
        return render_template("csv_import.html", stage="upload",
                               import_fields=CSV_IMPORT_FIELDS)

    # ── Step 1b: sheet selected from an Excel file ───────────────────────────
    if request.form.get("stage") == "sheet":
        wb_b64 = request.form.get("wb_b64", "")
        sheet_name = request.form.get("sheet_name", "")
        filename = request.form.get("filename", "")
        try:
            header_row = max(1, int(request.form.get("header_row", "1") or "1"))
        except ValueError:
            header_row = 1
        if not wb_b64 or not sheet_name:
            flash("Sheet selection is missing. Please upload the file again.", "danger")
            return redirect(url_for("import_csv"))
        try:
            workbook_bytes = base64.b64decode(wb_b64)
            content = _excel_sheet_to_csv(workbook_bytes, sheet_name, start_row=header_row)
        except Exception as exc:
            flash(f"Could not read sheet '{sheet_name}': {exc}", "danger")
            return redirect(url_for("import_csv"))

        reader = csv.reader(io.StringIO(content))
        headers = next(reader, [])
        if not headers:
            flash("The selected sheet has no header row.", "danger")
            return redirect(url_for("import_csv"))

        preview_rows = [row for i, row in enumerate(reader) if i < 5]

        field_keys = {f[0]: f[1] for f in CSV_IMPORT_FIELDS if f[0]}
        guessed = []
        for h in headers:
            normalised = h.lower().replace(" ", "_").replace("-", "_")
            match = ""
            for key in field_keys:
                if key in normalised or normalised in key:
                    match = key
                    break
            guessed.append(match)

        source_label = f"{filename} — {sheet_name}"
        if header_row > 1:
            source_label += f" (starting row {header_row})"
        return render_template(
            "csv_import.html",
            stage="map",
            headers=headers,
            preview_rows=preview_rows,
            guessed=guessed,
            csv_content=content,
            import_fields=CSV_IMPORT_FIELDS,
            source_label=source_label,
        )

    # ── Step 1: file just uploaded — detect type and show next step ──────────
    if "csv_file" in request.files and request.files["csv_file"].filename:
        file = request.files["csv_file"]
        filename = file.filename or ""
        ext = os.path.splitext(filename)[1].lower()
        raw_bytes = file.read()  # read once; file pointer is consumed after this

        # ── Excel path ───────────────────────────────────────────────────────
        if ext in _EXCEL_EXTENSIONS:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            except Exception as exc:
                flash(f"Could not open the Excel file: {exc}", "danger")
                return redirect(url_for("import_csv"))

            wb_b64 = base64.b64encode(raw_bytes).decode("ascii")
            return render_template(
                "csv_import.html",
                stage="sheet",
                sheet_names=sheet_names,
                wb_b64=wb_b64,
                filename=filename,
                import_fields=CSV_IMPORT_FIELDS,
            )

        # ── CSV path ─────────────────────────────────────────────────────────
        try:
            content = raw_bytes.decode("utf-8-sig")  # strip BOM if present
        except UnicodeDecodeError:
            content = raw_bytes.decode("latin-1")

        try:
            header_row = max(1, int(request.form.get("header_row", "1") or "1"))
        except ValueError:
            header_row = 1

        reader = csv.reader(io.StringIO(content))
        # Skip rows before the chosen header row.
        for _ in range(header_row - 1):
            next(reader, None)
        headers = next(reader, [])
        if not headers:
            flash("The uploaded file has no header row at the specified row.", "danger")
            return redirect(url_for("import_csv"))

        # Rebuild csv_content starting from the header row so the import stage
        # can read it as a plain header + data CSV without further offsets.
        remaining_rows = list(reader)
        csv_buf = io.StringIO()
        csv_writer = csv.writer(csv_buf)
        csv_writer.writerow(headers)
        csv_writer.writerows(remaining_rows)
        content = csv_buf.getvalue()

        # Read up to 5 preview rows.
        preview_rows = remaining_rows[:5]

        # Auto-guess mappings: compare lowercased CSV headers to field keys.
        field_keys = {f[0]: f[1] for f in CSV_IMPORT_FIELDS if f[0]}
        guessed = []
        for h in headers:
            normalised = h.lower().replace(" ", "_").replace("-", "_")
            match = ""
            for key in field_keys:
                if key in normalised or normalised in key:
                    match = key
                    break
            guessed.append(match)

        source_label = filename
        if header_row > 1:
            source_label += f" (starting row {header_row})"
        return render_template(
            "csv_import.html",
            stage="map",
            headers=headers,
            preview_rows=preview_rows,
            guessed=guessed,
            csv_content=content,
            import_fields=CSV_IMPORT_FIELDS,
            source_label=source_label,
        )

    # ── Step 2: column mapping submitted — run import ────────────────────────
    if request.form.get("stage") == "import":
        content = request.form.get("csv_content", "")
        reader = csv.reader(io.StringIO(content))
        headers = next(reader, [])

        # Build index→field map from the form (col_map_0, col_map_1 …).
        col_map = {}
        for i in range(len(headers)):
            field = request.form.get(f"col_map_{i}", "")
            if field:
                col_map[i] = field

        rows_to_import = []
        for raw_row in reader:
            record = {}
            for idx, field in col_map.items():
                if idx < len(raw_row):
                    record[field] = raw_row[idx].strip()
            rows_to_import.append(record)

        result = db.bulk_import_applications(rows_to_import)
        dup_note = (
            f", {result['duplicates']} duplicate(s) skipped" if result["duplicates"] else ""
        )
        flash(
            f"Import complete — {result['imported']} added, "
            f"{result['skipped']} skipped{dup_note}.",
            "success" if not result["errors"] else "warning",
        )
        return render_template(
            "csv_import.html",
            stage="result",
            result=result,
            import_fields=CSV_IMPORT_FIELDS,
        )

    flash("No file was uploaded.", "warning")
    return redirect(url_for("import_csv"))


# ---------------------------------------------------------------------------
# Companies
# ---------------------------------------------------------------------------

def _company_view_context():
    """Return (user_id, pool_enabled) for the current request."""
    login_enabled = db.get_setting("login_enabled", "0") == "1"
    pool_enabled = db.get_setting("company_pool_enabled", "0") == "1"
    user_id = session.get("user_id") if login_enabled else None
    return user_id, pool_enabled


@app.route("/companies")
@login_required
def companies():
    user_id, pool_enabled = _company_view_context()
    companies_list = db.get_companies(user_id=user_id, pool_enabled=pool_enabled)
    sector_freq = db.get_company_note_frequency()
    return render_template(
        "companies.html",
        companies=companies_list,
        years=db.YEARS,
        sector_freq=sector_freq,
        pool_enabled=pool_enabled,
        current_user_id=user_id,
    )


@app.route("/company/add", methods=["GET", "POST"])
@login_required
def add_company():
    if request.method == "POST":
        login_enabled = db.get_setting("login_enabled", "0") == "1"
        user_id = session.get("user_id") if login_enabled else None
        db.add_company(request.form, user_id=user_id)
        flash("Company added successfully.", "success")
        return redirect(url_for("companies"))
    return render_template(
        "company_form.html", company=None, years=db.YEARS, action="Add"
    )


@app.route("/company/edit/<int:company_id>", methods=["GET", "POST"])
@login_required
def edit_company(company_id):
    company = db.get_company(company_id)
    if not company:
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))
    if request.method == "POST":
        db.update_company(company_id, request.form)
        flash("Company updated.", "success")
        return redirect(url_for("companies"))
    return render_template(
        "company_form.html", company=company, years=db.YEARS, action="Edit"
    )


@app.route("/company/delete/<int:company_id>", methods=["POST"])
@login_required
def delete_company(company_id):
    db.delete_company(company_id)
    flash("Company deleted.", "warning")
    return redirect(url_for("companies"))


@app.route("/companies/bulk-delete", methods=["POST"])
@login_required
def bulk_delete_companies():
    raw_ids = request.form.getlist("selected_ids")
    # Accept only numeric strings to avoid injection; convert to int safely.
    selected_ids = []
    for x in raw_ids:
        try:
            n = int(x)
            if n > 0:
                selected_ids.append(n)
        except (ValueError, TypeError):
            pass
    if not selected_ids:
        flash("No companies selected.", "warning")
        return redirect(url_for("companies"))
    count = db.bulk_delete_companies(selected_ids)
    flash(f"Deleted {count} company record(s).", "warning")
    return redirect(url_for("companies"))


# ---------------------------------------------------------------------------
# Inbox (reminders)
# ---------------------------------------------------------------------------

@app.route("/inbox")
def inbox():
    reminders = db.get_reminders()
    return render_template("inbox.html", reminders=reminders)


@app.route("/inbox/dismiss/<int:reminder_id>", methods=["POST"])
def dismiss_reminder(reminder_id):
    db.dismiss_reminder(reminder_id)
    return redirect(url_for("inbox"))


@app.route("/inbox/dismiss-all", methods=["POST"])
def dismiss_all_reminders():
    db.dismiss_all_reminders()
    flash("All reminders dismissed.", "success")
    return redirect(url_for("inbox"))


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    section = request.args.get("section", "general")

    if request.method == "POST":
        action = request.form.get("action", "save_general")

        if action == "save_general":
            db.set_setting("reminder_enabled", "1" if request.form.get("reminder_enabled") else "0")
            days = request.form.get("reminder_days", "3").strip()
            if days.isdigit() and int(days) >= 1:
                db.set_setting("reminder_days", days)
            else:
                flash("Reminder days must be a positive integer.", "danger")
                return redirect(url_for("settings", section="general"))
            db.set_setting(
                "company_pool_enabled",
                "1" if request.form.get("company_pool_enabled") else "0",
            )
            flash("General settings saved.", "success")
            return redirect(url_for("settings", section="general"))

        elif action == "save_security":
            if DEPLOYMENT_MODE == "local":
                flash("Login / multi-user settings are not available in local mode.", "warning")
                return redirect(url_for("settings", section="users"))
            login_enabled = "1" if request.form.get("login_enabled") else "0"
            # If turning on login, ensure at least one user exists.
            if login_enabled == "1" and db.count_users() == 0:
                flash("Cannot enable login — no users exist. Add a user first.", "danger")
                return redirect(url_for("settings", section="users"))
            db.set_setting("login_enabled", login_enabled)
            flash("Security settings saved.", "success")
            return redirect(url_for("settings", section="users"))

        elif action == "add_status":
            ok, msg = db.add_status(request.form.get("name", ""))
            flash(msg, "success" if ok else "danger")
            return redirect(url_for("settings", section="statuses"))

        elif action == "delete_status":
            ok, msg = db.delete_status(request.form.get("name", ""))
            flash(msg, "success" if ok else "danger")
            return redirect(url_for("settings", section="statuses"))

        elif action == "add_user":
            if DEPLOYMENT_MODE == "local":
                flash("User management is not available in local mode.", "warning")
                return redirect(url_for("settings", section="users"))
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            password2 = request.form.get("password2", "")
            is_admin = bool(request.form.get("is_admin"))
            if not username or not password:
                flash("Username and password are required.", "danger")
            elif password != password2:
                flash("Passwords do not match.", "danger")
            elif len(password) < 8:
                flash("Password must be at least 8 characters.", "danger")
            else:
                pw_hash = generate_password_hash(password)
                ok, msg = db.add_user(username, pw_hash, is_admin)
                flash(msg, "success" if ok else "danger")
            return redirect(url_for("settings", section="users"))

        elif action == "delete_user":
            if DEPLOYMENT_MODE == "local":
                flash("User management is not available in local mode.", "warning")
                return redirect(url_for("settings", section="users"))
            user_id = request.form.get("user_id", "")
            if user_id.isdigit():
                ok, msg = db.delete_user(int(user_id))
                flash(msg, "success" if ok else "danger")
            return redirect(url_for("settings", section="users"))

        elif action == "save_ai":
            if DEPLOYMENT_MODE == "local":
                flash("AI settings are not available in local mode.", "warning")
                return redirect(url_for("settings", section="ai"))
            db.set_setting("ollama_enabled", "1" if request.form.get("ollama_enabled") else "0")
            db.set_setting("ai_fit_enabled", "1" if request.form.get("ai_fit_enabled") else "0")
            ollama_url = request.form.get("ollama_url", "").strip()
            if ollama_url:
                db.set_setting("ollama_url", ollama_url)
            db.set_setting("ollama_model", request.form.get("ollama_model", "llama3").strip() or "llama3")
            flash("AI settings saved.", "success")
            return redirect(url_for("settings", section="ai"))

        elif action == "save_profile":
            if DEPLOYMENT_MODE == "local":
                flash("AI profile settings are not available in local mode.", "warning")
                return redirect(url_for("settings", section="ai"))
            db.set_setting("user_profile_skills",     request.form.get("user_profile_skills",     "").strip())
            db.set_setting("user_profile_experience", request.form.get("user_profile_experience", "").strip())
            db.set_setting("user_profile_summary",    request.form.get("user_profile_summary",    "").strip())
            flash("Your profile has been saved.", "success")
            return redirect(url_for("settings", section="ai"))

        flash("Unknown action.", "warning")
        return redirect(url_for("settings", section=section))

    current = db.get_all_settings()
    statuses = db.get_status_options()
    users = db.get_users()
    return render_template(
        "settings.html",
        settings=current,
        section=section,
        statuses=statuses,
        users=users,
        app_version=APP_VERSION,
    )


@app.route("/settings/ollama-test", methods=["POST"])
@login_required
def ollama_test():
    """AJAX: test connection to the configured Ollama server."""
    ollama_url = request.json.get("url", db.get_setting("ollama_url", "http://localhost:11434"))
    ollama_url = ollama_url.rstrip("/")
    try:
        req = urllib.request.Request(
            f"{ollama_url}/api/tags",
            headers={"Accept": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
        models = [m.get("name", "") for m in data.get("models", [])]
        return jsonify({"ok": True, "models": models})
    except urllib.error.URLError:
        # Never expose internal URL, path, or OS details to the client.
        return jsonify({"ok": False, "error": "Could not connect to Ollama server. Check the URL and ensure the server is running."})
    except Exception:
        return jsonify({"ok": False, "error": "Could not connect to Ollama server."})


@app.route("/api/ollama-status")
@login_required
def ollama_status():
    """Lightweight ping: returns whether the configured Ollama server is reachable."""
    if db.get_setting("ollama_enabled", "0") != "1":
        return jsonify({"ok": False, "error": "Ollama is not enabled."})
    ollama_url = db.get_setting("ollama_url", "http://localhost:11434").rstrip("/")
    model = db.get_setting("ollama_model", "llama3")
    try:
        req = urllib.request.Request(
            f"{ollama_url}/api/tags",
            headers={"Accept": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=4) as resp:
            data = json.loads(resp.read().decode())
        models = [m.get("name", "") for m in data.get("models", [])]
        return jsonify({"ok": True, "model": model, "models": models})
    except urllib.error.URLError:
        return jsonify({"ok": False, "error": "Server unreachable."})
    except Exception:
        return jsonify({"ok": False, "error": "Could not connect to Ollama server."})


@app.route("/api/ai-fill", methods=["POST"])
@login_required
def ai_fill():
    """AJAX: send a pasted job description to Ollama and return extracted form fields."""
    import re as _re

    if db.get_setting("ollama_enabled", "0") != "1":
        return jsonify({"ok": False, "error": "Ollama AI Assistant is not enabled in Settings."})

    body = request.get_json(silent=True) or {}
    job_description = (body.get("job_description") or "").strip()
    if not job_description:
        return jsonify({"ok": False, "error": "Please paste a job description first."})

    ollama_url   = db.get_setting("ollama_url",   "http://localhost:11434").rstrip("/")
    ollama_model = db.get_setting("ollama_model", "llama3")

    prompt = (
        "You are a helpful assistant that extracts structured information from job postings.\n\n"
        "Given the job description below, return ONLY a single valid JSON object "
        "(no markdown fences, no extra text) with these keys:\n\n"
        '  "job_desc"  — the job title or role name (string)\n'
        '  "company"   — the company or organisation name (string)\n'
        '  "team"      — team, division, or department name, or "" if not stated (string)\n'
        '  "link"      — the application or posting URL if explicitly mentioned, or "" (string)\n'
        '  "comment"   — a concise 2–3 sentence summary of key requirements and responsibilities (string)\n\n'
        "Job Description:\n"
        "---\n"
        f"{job_description[:_MAX_JOB_DESC_LENGTH]}\n"   # cap to avoid huge prompts
        "---\n\n"
        "JSON object:"
    )

    payload = json.dumps({
        "model":  ollama_model,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.1},   # low temp → more deterministic JSON
    }).encode()

    try:
        req = urllib.request.Request(
            f"{ollama_url}/api/generate",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=90) as resp:
            result = json.loads(resp.read().decode())

        raw = (result.get("response") or "").strip()

        # Strip markdown code fences if the model wrapped the output anyway.
        fence_match = _re.search(r"```(?:json)?\s*([\s\S]+?)\s*```", raw)
        if fence_match:
            raw = fence_match.group(1).strip()

        # Extract the first JSON object in the response (handles preamble text).
        obj_match = _re.search(r"\{[\s\S]+\}", raw)
        if obj_match:
            raw = obj_match.group(0)

        fields = json.loads(raw)

        # Allow only expected keys; coerce values to strings.
        allowed = {"job_desc", "company", "team", "link", "comment"}
        fields = {k: str(v).strip() for k, v in fields.items() if k in allowed}

        return jsonify({"ok": True, "fields": fields})

    except urllib.error.URLError:
        return jsonify({"ok": False, "error": "Could not connect to the Ollama server. Is it running?"})
    except json.JSONDecodeError:
        return jsonify({
            "ok": False,
            "error": (
                "The AI returned an unrecognised format. "
                "Try a different model or simplify the job description."
            ),
        })
    except Exception:
        return jsonify({"ok": False, "error": "An unexpected error occurred. Please try again."})


@app.route("/api/upload-profile-pdf", methods=["POST"])
@login_required
def upload_profile_pdf():
    """AJAX: accept a PDF upload, extract its text, and store it as the user profile summary."""
    if db.get_setting("ollama_enabled", "0") != "1":
        return jsonify({"ok": False, "error": "Ollama AI Assistant is not enabled in Settings."})

    pdf_file = request.files.get("pdf")
    if not pdf_file or not pdf_file.filename:
        return jsonify({"ok": False, "error": "No file received."})

    filename_lower = pdf_file.filename.lower()
    if not filename_lower.endswith(".pdf"):
        return jsonify({"ok": False, "error": "Only PDF files are supported."})

    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(pdf_file.read()))
        pages_text = []
        for page in reader.pages:
            text = page.extract_text() or ""
            if text.strip():
                pages_text.append(text.strip())
        extracted = "\n\n".join(pages_text)
    except Exception:
        return jsonify({"ok": False, "error": "Could not read the PDF. Ensure it contains selectable text."})

    if not extracted.strip():
        return jsonify({"ok": False, "error": "No readable text found in the PDF. Try a text-based PDF."})

    # Truncate to a reasonable size before storing.
    extracted = extracted[:_MAX_PROFILE_LENGTH]

    db.set_setting("user_profile_summary", extracted)
    preview = extracted[:300].replace("\n", " ")
    return jsonify({"ok": True, "preview": preview, "char_count": len(extracted)})


@app.route("/api/ai-fit", methods=["POST"])
@login_required
def ai_fit():
    """AJAX: compare user profile with a job description and return a fit analysis."""
    import re as _re

    if db.get_setting("ollama_enabled", "0") != "1":
        return jsonify({"ok": False, "error": "Ollama AI Assistant is not enabled in Settings."})

    if db.get_setting("ai_fit_enabled", "0") != "1":
        return jsonify({"ok": False, "error": "Smart Job Fit Analysis is not enabled in Settings."})

    body = request.get_json(silent=True) or {}
    job_description = (body.get("job_description") or "").strip()
    if not job_description:
        return jsonify({"ok": False, "error": "No job description provided for fit analysis."})

    # Build user profile from stored settings.
    skills     = db.get_setting("user_profile_skills",     "").strip()
    experience = db.get_setting("user_profile_experience", "").strip()
    summary    = db.get_setting("user_profile_summary",    "").strip()

    if not (skills or experience or summary):
        return jsonify({
            "ok": False,
            "error": "Your profile is not set up yet. Go to Settings → AI Assistant and fill in your profile.",
        })

    profile_parts = []
    if summary:
        profile_parts.append(f"About me:\n{summary[:2000]}")
    if skills:
        profile_parts.append(f"My skills:\n{skills[:1000]}")
    if experience:
        profile_parts.append(f"My experience:\n{experience[:1500]}")
    profile_text = "\n\n".join(profile_parts)

    ollama_url   = db.get_setting("ollama_url",   "http://localhost:11434").rstrip("/")
    ollama_model = db.get_setting("ollama_model", "llama3")

    prompt = (
        "You are a career advisor. Given a candidate profile and a job description, "
        "evaluate how well the candidate fits the role.\n\n"
        "Return ONLY a single valid JSON object (no markdown fences, no extra text) with these keys:\n\n"
        '  "fit_score"        — integer 0–100 representing overall fit percentage\n'
        '  "verdict"          — one of: "Strong Fit", "Good Fit", "Moderate Fit", "Weak Fit", "Not a Fit"\n'
        '  "matching_skills"  — list of skills/qualities the candidate has that match the role (list of strings, max 6)\n'
        '  "skill_gaps"       — list of key requirements the candidate appears to lack (list of strings, max 5)\n'
        '  "recommendation"   — 2–3 sentence personalised recommendation for the candidate (string)\n\n'
        "Candidate Profile:\n"
        "---\n"
        f"{profile_text}\n"
        "---\n\n"
        "Job Description:\n"
        "---\n"
        f"{job_description[:_MAX_JOB_DESC_LENGTH]}\n"
        "---\n\n"
        "JSON object:"
    )

    payload = json.dumps({
        "model":   ollama_model,
        "prompt":  prompt,
        "stream":  False,
        "options": {"temperature": 0.2},
    }).encode()

    try:
        req = urllib.request.Request(
            f"{ollama_url}/api/generate",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode())

        raw = (result.get("response") or "").strip()

        # Strip markdown fences if present.
        fence_match = _re.search(r"```(?:json)?\s*([\s\S]+?)\s*```", raw)
        if fence_match:
            raw = fence_match.group(1).strip()

        # Extract first JSON object.
        obj_match = _re.search(r"\{[\s\S]+\}", raw)
        if obj_match:
            raw = obj_match.group(0)

        analysis = json.loads(raw)

        # Validate and sanitise the expected keys.
        _VERDICT_VALUES = {"Strong Fit", "Good Fit", "Moderate Fit", "Weak Fit", "Not a Fit"}
        fit_score = analysis.get("fit_score", 0)
        try:
            fit_score = max(0, min(100, int(fit_score)))
        except (TypeError, ValueError):
            fit_score = 0

        verdict = str(analysis.get("verdict", "Moderate Fit"))
        if verdict not in _VERDICT_VALUES:
            verdict = "Moderate Fit"

        matching_skills = [str(s)[:100] for s in (analysis.get("matching_skills") or [])[:6]]
        skill_gaps      = [str(s)[:100] for s in (analysis.get("skill_gaps")      or [])[:5]]
        recommendation  = str(analysis.get("recommendation", ""))[:600]

        return jsonify({
            "ok": True,
            "fit_score":       fit_score,
            "verdict":         verdict,
            "matching_skills": matching_skills,
            "skill_gaps":      skill_gaps,
            "recommendation":  recommendation,
        })

    except urllib.error.URLError:
        return jsonify({"ok": False, "error": "Could not connect to the Ollama server. Is it running?"})
    except json.JSONDecodeError:
        return jsonify({
            "ok": False,
            "error": "The AI returned an unrecognised format. Try a different model.",
        })
    except Exception:
        return jsonify({"ok": False, "error": "An unexpected error occurred during fit analysis."})


@app.route("/settings/check-update")
@login_required
def check_update():
    """AJAX: check GitHub releases for a newer version."""
    try:
        req = urllib.request.Request(
            f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest",
            headers={
                "Accept": "application/vnd.github+json",
                "User-Agent": "Job-Tracker-App",
            },
        )
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode())
        latest_tag = data.get("tag_name", "").lstrip("v")
        html_url = data.get("html_url", f"https://github.com/{GITHUB_REPO}/releases")
        is_newer = _version_is_newer(latest_tag, APP_VERSION)
        return jsonify({
            "ok": True,
            "current": APP_VERSION,
            "latest": latest_tag,
            "update_available": is_newer,
            "html_url": html_url,
        })
    except Exception:
        return jsonify({"ok": False, "error": "Could not reach GitHub to check for updates.", "current": APP_VERSION})


def _version_is_newer(latest: str, current: str) -> bool:
    """Return True if latest > current using simple numeric comparison."""
    try:
        l_parts = [int(x) for x in latest.split(".")]
        c_parts = [int(x) for x in current.split(".")]
        # Pad to same length.
        while len(l_parts) < len(c_parts):
            l_parts.append(0)
        while len(c_parts) < len(l_parts):
            c_parts.append(0)
        return l_parts > c_parts
    except Exception:
        return False


# Keep the old /statuses route working as a redirect to settings for backward compat.
@app.route("/statuses", methods=["GET", "POST"])
@login_required
def manage_statuses():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            ok, msg = db.add_status(request.form.get("name", ""))
            flash(msg, "success" if ok else "danger")
        elif action == "delete":
            ok, msg = db.delete_status(request.form.get("name", ""))
            flash(msg, "success" if ok else "danger")
    return redirect(url_for("settings", section="statuses"))


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@app.route("/export")
def export_page():
    status_options = db.get_status_options()
    return render_template(
        "export.html",
        years=db.YEARS,
        status_options=status_options,
    )


@app.route("/export/applications")
def export_applications():
    year = request.args.get("year", "")
    status = request.args.get("status", "")
    company = request.args.get("company", "").strip()

    apps = db.get_applications(
        year=int(year) if year.isdigit() else None,
        status=status if status else None,
    )
    if company:
        apps = [a for a in apps if company.lower() in a["company"].lower()]

    fields = [
        "id", "company", "job_desc", "team", "date_applied", "status",
        "cover_letter", "resume", "duration", "success_chance",
        "link", "contact", "comment", "additional_notes",
    ]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(apps)

    filename_parts = ["applications"]
    if year:
        filename_parts.append(year)
    if status:
        filename_parts.append(status)
    filename = "_".join(filename_parts) + ".csv"

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/export/companies")
def export_companies():
    companies_list = db.get_companies()
    fields = ["id", "company_name", "note",
              "applied_2023", "applied_2024", "applied_2025",
              "applied_2026", "applied_2027"]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(companies_list)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=companies.csv"},
    )


@app.route("/export/db")
def export_db():
    """Download a copy of the SQLite database for migration / backup."""
    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp.close()
    shutil.copy2(db.DB_PATH, tmp.name)
    return send_file(
        tmp.name,
        mimetype="application/octet-stream",
        as_attachment=True,
        download_name="jobs_backup.db",
    )


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=debug_mode, host="0.0.0.0", port=port)
