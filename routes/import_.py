"""CSV / Excel bulk import route."""

import csv
import io
import os
import uuid

from flask import (
    Blueprint, flash, redirect, render_template,
    request, session, url_for,
)
import openpyxl

import db
from .auth import login_required, current_user_id

bp = Blueprint("import_", __name__)

_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}

# Use a project-local directory instead of /tmp to avoid path restrictions.
_UPLOAD_TMP_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "_upload_tmp")
os.makedirs(_UPLOAD_TMP_DIR, exist_ok=True)

CSV_IMPORT_FIELDS = [
    ("",                "(ignore this column)"),
    ("company",         "Company *"),
    ("date_applied",    "Date Applied *"),
    ("job_desc",        "Job Description / Role"),
    ("team",            "Team / Division"),
    ("status",          "Status"),
    ("cover_letter",    "Cover Letter (1/0)"),
    ("resume",          "Resume (1/0)"),
    ("success_chance",  "Success Chance"),
    ("link",            "Application Link"),
    ("contact",         "Known Contact / Connection"),
    ("comment",         "Comment / Notes"),
    ("additional_notes", "Additional Notes"),
    ("industry",        "Industry / Sector"),
]


def _save_upload_tmp(data: bytes) -> str:
    path = os.path.join(_UPLOAD_TMP_DIR, f"{uuid.uuid4().hex}.bin")
    fd = os.open(path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
    try:
        with os.fdopen(fd, "wb") as fh:
            fh.write(data)
    except Exception:
        _delete_upload_tmp(path)
        raise
    return path


def _load_upload_tmp(path: str) -> bytes:
    with open(path, "rb") as fh:
        return fh.read()


def _delete_upload_tmp(path: str) -> None:
    try:
        os.remove(path)
    except OSError:
        pass


def _excel_sheet_to_csv(workbook_bytes: bytes, sheet_name: str, start_row: int = 1) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name]
    out = io.StringIO()
    writer = csv.writer(out)
    for row in ws.iter_rows(min_row=max(1, start_row), values_only=True):
        writer.writerow([("" if cell is None else str(cell)) for cell in row])
    wb.close()
    return out.getvalue()


@bp.route("/application/import", methods=["GET", "POST"])
@login_required
def import_csv():
    if request.method == "GET":
        return render_template("csv_import.html", stage="upload",
                               import_fields=CSV_IMPORT_FIELDS)

    if request.form.get("stage") == "sheet":
        sheet_name = request.form.get("sheet_name", "")
        wb_info = session.get("import_wb")
        if not wb_info or not sheet_name:
            flash("Sheet selection is missing. Please upload the file again.", "danger")
            return redirect(url_for("import_.import_csv"))
        filename = wb_info.get("filename", "")
        wb_path = wb_info.get("path", "")
        try:
            header_row = max(1, int(request.form.get("header_row", "1") or "1"))
        except ValueError:
            header_row = 1
        try:
            workbook_bytes = _load_upload_tmp(wb_path)
            content = _excel_sheet_to_csv(workbook_bytes, sheet_name, start_row=header_row)
        except Exception as exc:
            flash(f"Could not read sheet '{sheet_name}': {exc}", "danger")
            return redirect(url_for("import_.import_csv"))
        finally:
            _delete_upload_tmp(wb_path)
            session.pop("import_wb", None)

        reader = csv.reader(io.StringIO(content))
        headers = next(reader, [])
        if not headers:
            flash("The selected sheet has no header row.", "danger")
            return redirect(url_for("import_.import_csv"))

        preview_rows = [row for i, row in enumerate(reader) if i < 10]
        field_keys = {f[0]: f[1] for f in CSV_IMPORT_FIELDS if f[0]}
        guessed = []
        for h in headers:
            normalised = h.lower().replace(" ", "_").replace("-", "_")
            match = ""
            for key in field_keys:
                if key in normalised or normalised in key:
                    match = key
                    break
            guessed.append(match)

        source_label = f"{filename} — {sheet_name}"
        if header_row > 1:
            source_label += f" (starting row {header_row})"
        return render_template(
            "csv_import.html",
            stage="map",
            headers=headers,
            preview_rows=preview_rows,
            guessed=guessed,
            csv_content=content,
            import_fields=CSV_IMPORT_FIELDS,
            source_label=source_label,
        )

    if "csv_file" in request.files and request.files["csv_file"].filename:
        file = request.files["csv_file"]
        filename = file.filename or ""
        ext = os.path.splitext(filename)[1].lower()
        raw_bytes = file.read()

        if ext in _EXCEL_EXTENSIONS:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            except Exception as exc:
                flash(f"Could not open the Excel file: {exc}", "danger")
                return redirect(url_for("import_.import_csv"))

            wb_path = _save_upload_tmp(raw_bytes)
            session["import_wb"] = {"path": wb_path, "filename": filename}
            return render_template(
                "csv_import.html",
                stage="sheet",
                sheet_names=sheet_names,
                filename=filename,
                import_fields=CSV_IMPORT_FIELDS,
            )

        try:
            content = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            content = raw_bytes.decode("latin-1")

        try:
            header_row = max(1, int(request.form.get("header_row", "1") or "1"))
        except ValueError:
            header_row = 1

        reader = csv.reader(io.StringIO(content))
        for _ in range(header_row - 1):
            next(reader, None)
        headers = next(reader, [])
        if not headers:
            flash("The uploaded file has no header row at the specified row.", "danger")
            return redirect(url_for("import_.import_csv"))

        remaining_rows = list(reader)
        csv_buf = io.StringIO()
        csv_writer = csv.writer(csv_buf)
        csv_writer.writerow(headers)
        csv_writer.writerows(remaining_rows)
        content = csv_buf.getvalue()

        preview_rows = remaining_rows[:10]
        field_keys = {f[0]: f[1] for f in CSV_IMPORT_FIELDS if f[0]}
        guessed = []
        for h in headers:
            normalised = h.lower().replace(" ", "_").replace("-", "_")
            match = ""
            for key in field_keys:
                if key in normalised or normalised in key:
                    match = key
                    break
            guessed.append(match)

        source_label = filename
        if header_row > 1:
            source_label += f" (starting row {header_row})"
        return render_template(
            "csv_import.html",
            stage="map",
            headers=headers,
            preview_rows=preview_rows,
            guessed=guessed,
            csv_content=content,
            import_fields=CSV_IMPORT_FIELDS,
            source_label=source_label,
        )

    if request.form.get("stage") == "import":
        content = request.form.get("csv_content", "")
        reader = csv.reader(io.StringIO(content))
        headers = next(reader, [])

        col_map = {}
        for i in range(len(headers)):
            field = request.form.get(f"col_map_{i}", "")
            if field:
                col_map[i] = field

        rows_to_import = []
        for raw_row in reader:
            record = {}
            for idx, field in col_map.items():
                if idx < len(raw_row):
                    record[field] = raw_row[idx].strip()
            rows_to_import.append(record)

        result = db.bulk_import_applications(rows_to_import, user_id=current_user_id())
        dup_note = (
            f", {result['duplicates']} duplicate(s) skipped" if result["duplicates"] else ""
        )
        flash(
            f"Import complete — {result['imported']} added, "
            f"{result['skipped']} skipped{dup_note}.",
            "success" if not result["errors"] else "warning",
        )
        return render_template(
            "csv_import.html",
            stage="result",
            result=result,
            import_fields=CSV_IMPORT_FIELDS,
        )

    flash("No file was uploaded.", "warning")
    return redirect(url_for("import_.import_csv"))
